"""Generador de archivos Excel para facturas"""
from typing import Optional, List
from models import Factura, FacturaCabecera, FacturaDetalle

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None


class ExcelGenerator:
    """Clase para generar archivos Excel a partir de facturas"""
    
    def __init__(self):
        if openpyxl is None:
            raise ImportError(
                "openpyxl no está instalado. "
                "Ejecuta: pip install openpyxl"
            )
    
    def generar_excel(self, factura: Factura, ruta_salida: str):
        """Genera un archivo Excel con dos hojas: Cabecera y Detalle (una factura)"""
        self.generar_excel_multiple([factura], ruta_salida)
    
    def generar_excel_multiple(self, facturas: List[Factura], ruta_salida: str):
        """Genera un archivo Excel con múltiples facturas"""
        wb = openpyxl.Workbook()
        
        # Eliminar hoja por defecto
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Crear hoja de cabecera (una fila por factura)
        self._crear_hoja_cabecera_multiple(wb, facturas)
        
        # Crear hoja de detalle (todas las filas de todas las facturas)
        self._crear_hoja_detalle_multiple(wb, facturas)
        
        # Guardar archivo
        wb.save(ruta_salida)
    
    def _crear_hoja_cabecera(self, workbook, cabecera: FacturaCabecera):
        """Crea la hoja de cabecera con los datos de la factura (una factura)"""
        from models import Factura
        self._crear_hoja_cabecera_multiple(workbook, [Factura(cabecera=cabecera)])
    
    def _crear_hoja_cabecera_multiple(self, workbook, facturas: List[Factura]):
        """Crea la hoja de cabecera con múltiples facturas (una fila por factura)"""
        ws = workbook.create_sheet("Cabecera")
        
        # Estilos
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Título
        ws['A1'] = f"INFORMACIÓN DE FACTURAS ({len(facturas)} factura(s))"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells(f'A1:{get_column_letter(len(facturas[0].cabecera.to_dict()) + 1)}1')
        ws.row_dimensions[1].height = 25
        
        if not facturas:
            return
        
        # Obtener todos los campos posibles de todas las facturas
        todos_los_campos = set()
        for factura in facturas:
            todos_los_campos.update(factura.cabecera.to_dict().keys())
        
        campos_ordenados = sorted(todos_los_campos)
        
        # Encabezados de columnas
        row = 3
        col = 1
        for campo in campos_ordenados:
            cell = ws.cell(row=row, column=col, value=campo)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            col += 1
        
        # Datos de cada factura (una fila por factura)
        row = 4
        for factura in facturas:
            datos = factura.cabecera.to_dict()
            col = 1
            for campo in campos_ordenados:
                valor = datos.get(campo, '')
                cell = ws.cell(row=row, column=col, value=valor)
                cell.border = border
                cell.alignment = Alignment(vertical='center', wrap_text=True)
                col += 1
            row += 1
        
        # Ajustar ancho de columnas
        for col in range(1, len(campos_ordenados) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20
    
    def _crear_hoja_detalle(self, workbook, detalle: list):
        """Crea la hoja de detalle con los items de la factura (una factura)"""
        from models import Factura
        self._crear_hoja_detalle_multiple(workbook, [Factura(detalle=detalle)])
    
    def _crear_hoja_detalle_multiple(self, workbook, facturas: List[Factura]):
        """Crea la hoja de detalle con items de múltiples facturas"""
        ws = workbook.create_sheet("Detalle")
        
        # Estilos
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Contar total de items
        total_items = sum(len(f.detalle) for f in facturas)
        
        # Título
        ws['A1'] = f"DETALLE DE ITEMS ({total_items} items de {len(facturas)} factura(s))"
        ws['A1'].font = Font(bold=True, size=14)
        
        # Obtener headers de la primera factura que tenga items
        headers = None
        for factura in facturas:
            if factura.detalle:
                headers = list(factura.detalle[0].to_dict().keys())
                break
        
        if not headers:
            ws['A3'] = "No se encontraron items en el detalle"
            ws['A3'].font = Font(italic=True, size=10)
            return
        
        # Agregar columna para número de factura
        headers = ['N° Factura'] + headers
        ws.merge_cells(f'A1:{get_column_letter(len(headers))}1')
        ws.row_dimensions[1].height = 25
        
        # Encabezados de columnas
        col = 1
        for header in headers:
            cell = ws.cell(row=3, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            col += 1
        
        # Datos de todas las facturas
        row = 4
        for factura_idx, factura in enumerate(facturas):
            num_factura = factura.cabecera.numero_factura or f"Factura {factura_idx + 1}"
            
            for item in factura.detalle:
                datos = item.to_dict()
                col = 1
                
                # Número de factura
                cell = ws.cell(row=row, column=col, value=num_factura)
                cell.border = border
                cell.alignment = Alignment(vertical='center')
                col += 1
                
                # Datos del item
                for header in headers[1:]:  # Saltar 'N° Factura'
                    valor = datos.get(header, '')
                    cell = ws.cell(row=row, column=col, value=valor)
                    cell.border = border
                    # Alinear números a la derecha
                    if isinstance(valor, (int, float)) or (isinstance(valor, str) and valor and valor.replace('.', '').replace(',', '').replace('-', '').isdigit()):
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                    else:
                        cell.alignment = Alignment(vertical='center', wrap_text=True)
                    col += 1
                row += 1
        
        # Ajustar ancho de columnas
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def generar_excel_pandas(self, factura: Factura, ruta_salida: str):
        """Alternativa usando pandas (más simple pero menos control de formato)"""
        try:
            import pandas as pd
        except ImportError:
            raise ImportError("pandas no está instalado. Ejecuta: pip install pandas")
        
        # Crear DataFrame para cabecera
        df_cabecera = pd.DataFrame([factura.cabecera.to_dict()])
        
        # Crear DataFrame para detalle
        if factura.detalle:
            df_detalle = pd.DataFrame([item.to_dict() for item in factura.detalle])
        else:
            df_detalle = pd.DataFrame()
        
        # Escribir a Excel con múltiples hojas
        with pd.ExcelWriter(ruta_salida, engine='openpyxl') as writer:
            df_cabecera.to_excel(writer, sheet_name='Cabecera', index=False)
            if not df_detalle.empty:
                df_detalle.to_excel(writer, sheet_name='Detalle', index=False)

